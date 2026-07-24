#!/usr/bin/env python3
"""Normalize IP register-map workbooks into CSV, JSON, and Markdown."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FIELDS = [
    "ip_core", "access", "address_raw", "address_value", "byte_offset",
    "address_scale", "name", "description", "notes", "source_workbook",
    "source_sheet", "source_row",
]
WRITE_OFFSET = "\u5199\u504f\u79fb\u5730\u5740"
READ_OFFSET = "\u8bfb\u504f\u79fb\u5730\u5740"
ADDRESS_PREFIX = re.compile(r"^\s*(0[xX][0-9A-Fa-f]+|[0-9]+)(?=$|[\s\uff08(])")
DEFAULT_BASE = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = DEFAULT_BASE / "origin_v3"
DEFAULT_OUTPUT = DEFAULT_BASE / "generated"


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return "\n".join(line.rstrip() for line in text.strip().split("\n"))


def parse_address_value(text: str) -> int | None:
    """Parse a leading integer while preserving width notes in address_raw."""
    match = ADDRESS_PREFIX.match(text)
    return int(match.group(1), 0) if match else None


def core_name(path: Path) -> str:
    name = path.stem.strip()
    suffix = "IP\u6838\u901a\u7528\u578b\u5730\u5740\u5206\u914d\u8868"
    name = re.sub(
        rf"\s*{suffix}(?:_\u5e26\u64cd\u4f5c)?(?:\uff08\u516c\u5f00\uff09)?\s*$",
        "",
        name,
    )
    return name.strip()


def select_workbooks(workbooks: list[Path]) -> list[Path]:
    """Choose one source per IP core, preferring the operation-annotated variant."""
    grouped: dict[str, list[Path]] = {}
    for workbook in workbooks:
        if workbook.name.startswith("~$"):
            continue
        grouped.setdefault(core_name(workbook), []).append(workbook)

    selected = []
    for name in sorted(grouped):
        candidates = grouped[name]
        best_priority = max("_\u5e26\u64cd\u4f5c" in item.stem for item in candidates)
        best = sorted(
            item for item in candidates
            if ("_\u5e26\u64cd\u4f5c" in item.stem) == best_priority
        )
        if len(best) != 1:
            paths = ", ".join(str(item) for item in best)
            raise ValueError(f"Multiple equally preferred workbooks for {name}: {paths}")
        selected.append(best[0])
    return selected


def merged_value(ws, row: int, column: int) -> str:
    if ws.cell(row, column).value is not None:
        return clean(ws.cell(row, column).value)
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return clean(ws.cell(merged.min_row, merged.min_col).value)
    return ""


def access_type(sheet: str, group: str, title: str) -> str:
    if WRITE_OFFSET in group:
        return "write"
    if READ_OFFSET in group:
        return "read"
    combined = f"{sheet}\n{title}"
    if "\u5199" in sheet or WRITE_OFFSET in combined:
        return "write"
    if "\u8bfb" in sheet or READ_OFFSET in combined:
        return "read"
    return "unknown"


def address_scale(group: str, title: str) -> int:
    return 4 if re.search("\u4e58\u4ee5\\s*4", f"{group}\n{title}") else 1


def parse_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    records = []
    for ws in workbook.worksheets:
        title = merged_value(ws, 1, 1)
        for row in range(3, ws.max_row + 1):
            raw = clean(ws.cell(row, 2).value)
            if not raw:
                continue
            value = parse_address_value(raw)
            if value is None:
                continue
            group = merged_value(ws, row, 1)
            scale = address_scale(group, title)
            access = access_type(ws.title, group, title)
            if access == "unknown":
                # “带操作”表下半部分是命令步骤，不是新的寄存器定义。
                continue
            records.append({
                "ip_core": core_name(path),
                "access": access,
                "address_raw": raw,
                "address_value": value,
                "byte_offset": value * scale,
                "address_scale": scale,
                "name": merged_value(ws, row, 3),
                "description": merged_value(ws, row, 4),
                "notes": merged_value(ws, row, 5) if ws.max_column >= 5 else "",
                "source_workbook": path.name,
                "source_sheet": ws.title,
                "source_row": row,
            })
    return records


def write_csv(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(records)


def write_json(path: Path, records: list[dict[str, Any]]) -> None:
    names = sorted({record["ip_core"] for record in records})
    payload = {
        "schema_version": 1,
        "address_rule": "byte_offset = address_value * address_scale; scale is 4 only when explicitly stated by the source.",
        "register_count": len(records),
        "ip_cores": [{
            "name": name,
            "register_count": sum(record["ip_core"] == name for record in records),
            "registers": [record for record in records if record["ip_core"] == name],
        } for name in names],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def md_cell(value: Any) -> str:
    return clean(value).replace("|", "\\|").replace("\n", "<br>")


def write_markdown(path: Path, records: list[dict[str, Any]]) -> None:
    counts = Counter(record["ip_core"] for record in records)
    lines = [
        "# IP \u6838\u5bc4\u5b58\u5668\u5730\u5740\u89c4\u8303\u5316\u6e05\u5355", "",
        "> \u672c\u6587\u4ef6\u7531 `tools/export_register_maps.py` \u4ece\u539f\u59cb XLSX \u81ea\u52a8\u751f\u6210\uff0c\u8bf7\u52ff\u624b\u5de5\u4fee\u6539\u3002", "",
        "## \u5730\u5740\u89c4\u5219", "",
        "- `address_raw`: \u6e90\u6587\u6863\u4e2d\u7684\u5730\u5740\u6587\u672c\u3002",
        "- `address_scale`: \u6e90\u6587\u6863\u660e\u786e\u8981\u6c42\u201c\u4e58\u4ee54\u201d\u65f6\u4e3a `4`\uff0c\u5426\u5219\u4e3a `1`\u3002",
        "- `byte_offset`: `address_value * address_scale`\u3002", "",
        "## \u6c47\u603b", "", "| IP \u6838 | \u5bc4\u5b58\u5668\u6570 |", "|---|---:|",
    ]
    lines.extend(f"| {md_cell(name)} | {counts[name]} |" for name in sorted(counts))
    for name in sorted(counts):
        lines += ["", f"## {name}", "",
                  "| \u8bbf\u95ee | \u539f\u5730\u5740 | \u5b57\u8282\u504f\u79fb | \u540d\u79f0 | \u529f\u80fd\u53ca\u4f7f\u7528\u65b9\u6cd5 | \u9644\u6ce8 | \u6765\u6e90 |",
                  "|---|---:|---:|---|---|---|---|"]
        for record in (item for item in records if item["ip_core"] == name):
            source = f'{record["source_workbook"]} / {record["source_sheet"]}:{record["source_row"]}'
            cells = [record["access"], f'`{record["address_raw"]}`',
                     f'`0x{record["byte_offset"]:X}`', record["name"],
                     record["description"], record["notes"], source]
            lines.append("| " + " | ".join(md_cell(cell) for cell in cells) + " |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def safe_name(name: str) -> str:
    return re.sub(r"[^0-9A-Za-z_.-]+", "_", name).strip("_")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    source = args.source.resolve()
    output = args.output.resolve()
    output.mkdir(parents=True, exist_ok=True)
    discovered = sorted(
        workbook for workbook in source.glob("*.xlsx")
        if not workbook.name.startswith("~$")
    )
    if not discovered:
        raise SystemExit(f"No XLSX files found in {source}")
    workbooks = select_workbooks(discovered)
    records = [record for workbook in workbooks for record in parse_workbook(workbook)]
    records.sort(key=lambda item: (item["ip_core"], item["source_sheet"], item["source_row"]))
    write_csv(output / "registers.csv", records)
    write_json(output / "registers.json", records)
    write_markdown(output / "registers.md", records)
    for name in sorted({record["ip_core"] for record in records}):
        write_csv(output / f"{safe_name(name)}.csv", [r for r in records if r["ip_core"] == name])
    print(
        f"Exported {len(records)} registers from {len(workbooks)} selected workbooks "
        f"({len(discovered)} discovered) to {output}"
    )


if __name__ == "__main__":
    main()
